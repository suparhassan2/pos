#!/usr/bin/env python3
"""
Position Reconciliation Tool: XTP vs JPM

This script reconciles position files between XTP and JPM systems by:
1. Loading position data from both systems
2. Applying strike and settlement price multipliers
3. Aggregating quantities by Account + Product Code + Adjusted Prices
4. Comparing matched and unmatched positions
5. Generating detailed reconciliation reports
"""

import pandas as pd
import numpy as np
import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional


class ReconciliationTool:
    """Main reconciliation tool class"""

    # Per-source column mapping configuration.
    # 'exact'              : {std_col: src_col}  — checked first; exact source column names
    # 'product_code_cols'  : [src_col, ...]      — priority list for Product_Code
    # 'flexible'           : {std_col: fn(col_lower)->bool}  — fallback pattern matching
    # 'product_code_flexible': fn(col_lower)->bool  — last-resort Product_Code match
    _SOURCE_CONFIGS: Dict = {
        'XTP': {
            # read_options are passed directly to pd.read_csv / pd.read_excel.
            # header=1 means row 2 in Excel (0-indexed) contains the column names.
            'read_options': {'header': 1},
            'exact': {
                'Account':          'JPM Account',
                'Strike_Price':     'Strike',
                'Settlement_Price': 'SettlementPrice',
            },
            'product_code_cols': ['JPM GMI Code', 'Clearing Code', 'Exchange Clearing Code'],
            'flexible': {
                'Account':          lambda c: 'account' in c,
                'Strike_Price':     lambda c: 'strike' in c,
                'Settlement_Price': lambda c: 'settle' in c,
                'Quantity':         lambda c: 'qty' in c or 'quantity' in c,
            },
            'product_code_flexible': lambda c: ('clearing' in c or 'product' in c) and 'code' in c,
            # Exclusion filters applied to raw source column names before standardization.
            # Each rule: {'column': str, 'test': 'blank'|'in'|'startswith'|'endswith'|'contains',
            #             'values': [str, ...]}   (values not needed for 'blank')
            'exclusions': [
                # Exclude rows where XTP Account starts with "HOC" or ends with "T"
                {'column': 'XTP Account', 'test': 'startswith', 'values': ['HOC']},
                {'column': 'XTP Account', 'test': 'endswith',   'values': ['T']},
                # Exclude rows where JPM Account is blank or is account 12345
                {'column': 'JPM Account', 'test': 'blank'},
                {'column': 'JPM Account', 'test': 'in',         'values': ['12345']},
                # Exclude rows where XTP Contract code contains CADBONDS
                {'column': 'XTP Contract code', 'test': 'contains', 'values': ['CADBONDS']},
            ],
        },
        'JPM': {
            'exact': {
                'Account':          'JPM ID / Broker ID',
                'Strike_Price':     'Strike',
                'Settlement_Price': 'Settlement_Price',
            },
            'product_code_cols': ['Product / Exchange Ticker'],
            # Row-level fallbacks: if Product_Code is blank after primary assignment,
            # try these columns in order until a non-blank value is found.
            'product_code_fallbacks': ['Clearing Code'],
            'flexible': {
                'Account':          lambda c: 'account' in c or ('jpm' in c and 'id' in c),
                'Strike_Price':     lambda c: 'strike' in c,
                'Settlement_Price': lambda c: 'settle' in c,
                'Quantity':         lambda c: 'qty' in c or 'quantity' in c,
            },
            'product_code_flexible': lambda c: 'product' in c and ('code' in c or 'ticker' in c),
            # Exclude specific JPM accounts
            'exclusions': [
                {'column': 'JPM ID / Broker ID', 'test': 'in',
                 'values': ['1111', '2222', '2321', '23122']},
            ],
        },
    }

    def __init__(self, config_dir: str = "config", log_dir: str = "logs",
                 strike_tolerance: float = 0.01, qty_tolerance: float = 0.0,
                 qty_percent_tolerance: float = 0.0):
        """
        Initialize the reconciliation tool.

        Args:
            config_dir: Directory containing multiplier CSV files
            log_dir: Directory for log files
            strike_tolerance: Reserved for future fuzzy strike price matching (currently unused)
            qty_tolerance: Absolute tolerance for quantity differences
            qty_percent_tolerance: Fractional tolerance for quantity differences (e.g. 0.05 = 5%)
        """
        if qty_tolerance < 0:
            raise ValueError(f"qty_tolerance must be non-negative, got {qty_tolerance}")
        if qty_percent_tolerance < 0:
            raise ValueError(f"qty_percent_tolerance must be non-negative, got {qty_percent_tolerance}")

        self.config_dir = Path(config_dir)
        self.log_dir = Path(log_dir)
        self.strike_tolerance = strike_tolerance  # reserved for future fuzzy strike matching
        self.qty_tolerance = qty_tolerance
        self.qty_percent_tolerance = qty_percent_tolerance

        self.config_dir.mkdir(parents=True, exist_ok=True)
        self.log_dir.mkdir(parents=True, exist_ok=True)

        self._setup_logging()

        # Multipliers: independent per source (XTP/JPM) and price type (strike/settle)
        self.multipliers: Dict = {
            'strike_xtp': {},
            'strike_jpm': {},
            'settle_xtp': {},
            'settle_jpm': {},
        }
        self.products_with_default_multipliers: set = set()
        self.logger.info("Reconciliation Tool initialized")

    def _setup_logging(self):
        """Setup a dedicated logger (does not modify the root logger)."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = self.log_dir / f"reconcile_{timestamp}.log"

        self.logger = logging.getLogger(f"{__name__}.{id(self)}")
        self.logger.setLevel(logging.INFO)
        self.logger.handlers.clear()  # avoid duplicate handlers on re-init

        fmt = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

        fh = logging.FileHandler(log_file)
        fh.setFormatter(fmt)
        sh = logging.StreamHandler(sys.stdout)
        sh.setFormatter(fmt)
        self.logger.addHandler(fh)
        self.logger.addHandler(sh)
        self.logger.info(f"Logging initialized: {log_file}")

    def load_multipliers(self, xtp_file: Optional[str] = None,
                         jpm_file: Optional[str] = None) -> Dict:
        """
        Load multipliers from two source-specific CSV files.

        Each file has the format:
            Product_Code, Strike_Multiplier, Settle_Multiplier

        Default file locations:
            XTP: config/xtp_multipliers.csv
            JPM: config/jpm_multipliers.csv

        A header-only template is created automatically for any file that does not exist.

        Returns:
            self.multipliers dict
        """
        xtp_path = Path(xtp_file) if xtp_file else self.config_dir / "xtp_multipliers.csv"
        jpm_path = Path(jpm_file) if jpm_file else self.config_dir / "jpm_multipliers.csv"

        for path, source in [(xtp_path, 'XTP'), (jpm_path, 'JPM')]:
            if not path.exists():
                self._create_template_multipliers(path, source)
                self.logger.warning(
                    f"No {source} multiplier data loaded — template created at {path}. "
                    "Fill it in with your product multipliers before running reconciliation."
                )
            else:
                self._load_source_multipliers(path, source)

        return self.multipliers

    def _create_template_multipliers(self, filepath: Path, source: str):
        """Create a header-only template multipliers CSV for a given source."""
        with open(filepath, 'w') as f:
            f.write("Product_Code,Strike_Multiplier,Settle_Multiplier\n")
            f.write("# Add one row per product. Use 1.0 if no scaling is needed.\n")
        self.logger.info(f"Template {source} multipliers file created: {filepath}")

    def _load_source_multipliers(self, filepath: Path, source: str):
        """
        Load a source-specific multipliers CSV.

        Expected format: Product_Code, Strike_Multiplier, Settle_Multiplier

        Args:
            filepath: Path to the multipliers CSV
            source: 'XTP' or 'JPM' — determines which internal keys are populated
        """
        df = pd.read_csv(filepath, comment='#')

        required = ['Product_Code', 'Strike_Multiplier', 'Settle_Multiplier']
        missing = [c for c in required if c not in df.columns]
        if missing:
            raise ValueError(
                f"{source} multipliers file '{filepath}' is missing columns: {missing}. "
                f"Expected: Product_Code, Strike_Multiplier, Settle_Multiplier. "
                f"Found: {list(df.columns)}"
            )

        products = df['Product_Code'].astype(str).str.upper().str.strip()
        strike_vals = pd.to_numeric(df['Strike_Multiplier'], errors='coerce')
        settle_vals = pd.to_numeric(df['Settle_Multiplier'], errors='coerce')

        for col_name, vals in [('Strike_Multiplier', strike_vals), ('Settle_Multiplier', settle_vals)]:
            invalid = vals.isna().sum()
            if invalid:
                self.logger.warning(
                    f"{source}: {invalid} non-numeric value(s) in '{col_name}' — defaulting to 1.0"
                )

        src = source.lower()
        self.multipliers[f'strike_{src}'] = dict(zip(products, strike_vals.fillna(1.0)))
        self.multipliers[f'settle_{src}'] = dict(zip(products, settle_vals.fillna(1.0)))
        self.logger.info(f"Loaded {source} multipliers from {filepath} — {len(products)} products")

    def get_multiplier(self, product_code: str, price_type: str, source: str) -> float:
        """
        Get multiplier for a product code.

        Args:
            product_code: Product code to look up
            price_type: 'strike' or 'settle'
            source: 'XTP' or 'JPM'

        Returns:
            Multiplier value (default 1.0 if not found)
        """
        product_code = str(product_code).upper().strip()
        key = f"{price_type}_{source.lower()}"
        multipliers = self.multipliers.get(key, {})
        if product_code not in multipliers:
            if product_code not in self.products_with_default_multipliers:
                self.logger.warning(
                    f"Product '{product_code}' not in {price_type} multipliers for {source}. "
                    "Using default 1.0"
                )
                self.products_with_default_multipliers.add(product_code)
            return 1.0
        return multipliers[product_code]

    def load_position_file(self, filepath: str, file_type: str) -> pd.DataFrame:
        """
        Load a position file (Excel or CSV).

        Args:
            filepath: Path to position file
            file_type: 'XTP' or 'JPM'

        Returns:
            DataFrame with position data
        """
        filepath = Path(filepath)
        if not filepath.exists():
            raise FileNotFoundError(f"{file_type} position file not found: {filepath}")

        read_opts = self._SOURCE_CONFIGS.get(file_type, {}).get('read_options', {})
        if read_opts:
            self.logger.info(f"Applying read options for {file_type}: {read_opts}")

        self.logger.info(f"Loading {file_type} positions from: {filepath}")
        try:
            if filepath.suffix.lower() == '.csv':
                df = pd.read_csv(filepath, **read_opts)
            elif filepath.suffix.lower() in ['.xlsx', '.xls']:
                df = pd.read_excel(filepath, **read_opts)
            else:
                raise ValueError(f"Unsupported file format: {filepath.suffix}")
            self.logger.info(f"Loaded {len(df)} records from {file_type} file")
            self.logger.info(f"Columns: {list(df.columns)}")
            return df
        except Exception as e:
            self.logger.error(f"Error loading {file_type} file: {e}")
            raise

    def _apply_exclusions(self, df: pd.DataFrame, source: str) -> pd.DataFrame:
        """
        Apply pre-standardization exclusion filters defined in _SOURCE_CONFIGS[source]['exclusions'].

        Each rule is a dict with:
            column : str   — raw source column name to filter on
            test   : str   — one of: 'blank', 'in', 'startswith', 'endswith', 'contains'
            values : list  — list of strings to match against (not required for 'blank')

        Matching is case-insensitive for 'contains'; case-sensitive for all others.
        A warning is logged for any rule whose column is absent from the file.
        """
        rules = self._SOURCE_CONFIGS[source].get('exclusions', [])
        if not rules:
            return df

        initial_len = len(df)

        for rule in rules:
            col = rule['column']
            test = rule['test']

            if col not in df.columns:
                self.logger.warning(
                    f"{source}: exclusion column '{col}' not found in file — rule skipped. "
                    f"Available columns: {list(df.columns)}"
                )
                continue

            before = len(df)

            if test == 'blank':
                # Keep rows where the column has a non-blank, non-null value
                raw = df[col]
                is_blank = raw.isna() | raw.astype(str).str.strip().isin(['', 'nan', 'NaN', 'None', 'none'])
                df = df[~is_blank].copy()

            else:
                values = [str(v).strip() for v in rule.get('values', [])]
                col_str = df[col].astype(str).str.strip()

                if test == 'in':
                    mask = ~col_str.isin(values)
                elif test == 'startswith':
                    mask = ~col_str.str.startswith(tuple(values))
                elif test == 'endswith':
                    mask = ~col_str.str.endswith(tuple(values))
                elif test == 'contains':
                    pattern = '|'.join(values)
                    mask = ~col_str.str.contains(pattern, case=False, na=False, regex=True)
                else:
                    self.logger.warning(f"{source}: unknown exclusion test '{test}' — rule skipped")
                    continue

                df = df[mask].copy()

            excluded = before - len(df)
            if excluded > 0:
                self.logger.info(
                    f"{source}: excluded {excluded} row(s) — '{col}' {test}"
                    + (f" {rule.get('values')}" if test != 'blank' else "")
                )

        total_excluded = initial_len - len(df)
        if total_excluded > 0:
            self.logger.info(
                f"{source}: {total_excluded} row(s) excluded by filters; {len(df)} remaining"
            )
        return df

    def _prepare_data(self, df: pd.DataFrame, source: str) -> pd.DataFrame:
        """
        Prepare position data for reconciliation: map columns, clean types,
        separate incomplete records, and apply multipliers.

        Column resolution order (highest to lowest priority):
          1. Exact source column name match  (cfg['exact'])
          2. Priority list for Product_Code  (cfg['product_code_cols'])
          3. Flexible pattern matching        (cfg['flexible'])
          4. Last-resort Product_Code pattern (cfg['product_code_flexible'])

        Args:
            df: Raw position DataFrame
            source: 'XTP' or 'JPM'

        Returns:
            Prepared DataFrame with standardized columns and _is_complete flag
        """
        self.logger.info(f"Preparing {source} data...")
        cfg = self._SOURCE_CONFIGS[source]
        df = df.copy()

        # Apply exclusion filters on raw column names before any standardization
        df = self._apply_exclusions(df, source)

        column_mapping: Dict[str, str] = {}
        already_mapped: set = set()

        # Step 1: Exact column name matches
        for std_col, src_col in cfg['exact'].items():
            if src_col in df.columns and src_col not in column_mapping:
                column_mapping[src_col] = std_col
                already_mapped.add(std_col)

        # Step 2: Product_Code — priority list then flexible
        product_code_source = None
        if 'Product_Code' not in df.columns:
            for src_col in cfg['product_code_cols']:
                if src_col in df.columns:
                    df['Product_Code'] = df[src_col]
                    product_code_source = src_col
                    break
            if 'Product_Code' not in df.columns:
                pc_flex = cfg.get('product_code_flexible')
                for col in df.columns:
                    if pc_flex and pc_flex(col.lower()):
                        df['Product_Code'] = df[col]
                        product_code_source = col
                        break

        if 'Product_Code' not in df.columns:
            raise ValueError(
                f"{source} file must contain a product code column "
                f"(tried: {cfg['product_code_cols']}). "
                f"Available columns: {list(df.columns)}"
            )

        # Step 2b: Row-level fallbacks — fill blank Product_Code values from fallback columns
        for fallback_col in cfg.get('product_code_fallbacks', []):
            if fallback_col not in df.columns:
                continue
            blank_mask = df['Product_Code'].isna() | (df['Product_Code'].astype(str).str.strip() == '')
            n_blank = blank_mask.sum()
            if n_blank == 0:
                break
            df.loc[blank_mask, 'Product_Code'] = df.loc[blank_mask, fallback_col]
            filled = blank_mask.sum() - (
                df['Product_Code'].isna() | (df['Product_Code'].astype(str).str.strip() == '')
            ).sum()
            self.logger.info(
                f"{source}: filled {filled} blank Product_Code row(s) from '{fallback_col}'"
            )

        # Step 3: Flexible matching for remaining standard columns
        needed = {'Account', 'Strike_Price', 'Settlement_Price', 'Quantity'} - already_mapped
        for col in df.columns:
            if not needed:
                break
            if col in column_mapping:
                continue
            col_lower = col.lower()
            for std_col, pattern_fn in cfg['flexible'].items():
                if std_col in needed and pattern_fn(col_lower):
                    column_mapping[col] = std_col
                    needed.discard(std_col)
                    break

        df = df.rename(columns=column_mapping)

        if column_mapping:
            self.logger.info(f"{source} column mappings: {column_mapping}")
        if product_code_source:
            self.logger.info(f"{source} Product_Code mapped from: {product_code_source}")

        required_cols = ['Account', 'Product_Code', 'Strike_Price', 'Settlement_Price', 'Quantity']
        missing_cols = [c for c in required_cols if c not in df.columns]
        if missing_cols:
            raise ValueError(
                f"{source} file missing required columns: {missing_cols}. "
                f"Available: {list(df.columns)}"
            )

        # Clean and convert types
        df['Account'] = df['Account'].astype(str).str.strip()
        df['Product_Code'] = df['Product_Code'].astype(str).str.strip().str.upper()
        df['Strike_Price'] = pd.to_numeric(df['Strike_Price'], errors='coerce')
        df['Settlement_Price'] = pd.to_numeric(df['Settlement_Price'], errors='coerce')
        df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce')

        # Mark completeness and split
        df['_is_complete'] = df[required_cols].notna().all(axis=1)
        df_complete = df[df['_is_complete']].copy()
        df_incomplete = df[~df['_is_complete']].copy()

        if len(df_incomplete) > 0:
            self.logger.info(f"Found {len(df_incomplete)} {source} records with missing data")

        # Apply multipliers to complete records only
        if len(df_complete) > 0:
            df_complete['Strike_Multiplier'] = df_complete['Product_Code'].apply(
                lambda x: self.get_multiplier(x, 'strike', source)
            )
            df_complete['Settle_Multiplier'] = df_complete['Product_Code'].apply(
                lambda x: self.get_multiplier(x, 'settle', source)
            )
            df_complete['Strike_Price_Adjusted'] = (
                df_complete['Strike_Price'] * df_complete['Strike_Multiplier']
            ).round(2)
            df_complete['Settlement_Price_Adjusted'] = (
                df_complete['Settlement_Price'] * df_complete['Settle_Multiplier']
            ).round(2)

        if len(df_incomplete) > 0:
            for col in ['Strike_Multiplier', 'Settle_Multiplier',
                        'Strike_Price_Adjusted', 'Settlement_Price_Adjusted']:
                df_incomplete[col] = None

        result = pd.concat([df_complete, df_incomplete], ignore_index=True)
        self.logger.info(
            f"Prepared {len(df_complete)} complete {source} records, "
            f"{len(df_incomplete)} incomplete"
        )
        return result

    def aggregate_positions(self, df: pd.DataFrame, source: str) -> pd.DataFrame:
        """Aggregate positions by Account + Product Code + Adjusted Prices."""
        self.logger.info(f"Aggregating {source} positions...")
        group_cols = ['Account', 'Product_Code', 'Strike_Price_Adjusted', 'Settlement_Price_Adjusted']
        agg_df = df.groupby(group_cols, as_index=False).agg({
            'Quantity': 'sum',
            'Strike_Price': 'first',
            'Settlement_Price': 'first',
            'Strike_Multiplier': 'first',
            'Settle_Multiplier': 'first',
        })
        self.logger.info(f"Aggregated {len(df)} {source} records into {len(agg_df)} groups")
        return agg_df

    def reconcile(self, xtp_file: str, jpm_file: str,
                  xtp_multipliers_file: Optional[str] = None,
                  jpm_multipliers_file: Optional[str] = None) -> Dict:
        """
        Perform reconciliation between XTP and JPM positions.

        Args:
            xtp_file: Path to XTP position file
            jpm_file: Path to JPM position file
            xtp_multipliers_file: Optional path to XTP multipliers CSV
                                  (default: config/xtp_multipliers.csv)
            jpm_multipliers_file: Optional path to JPM multipliers CSV
                                  (default: config/jpm_multipliers.csv)

        Returns:
            Dictionary containing reconciliation results
        """
        self.logger.info("=" * 80)
        self.logger.info("Starting Reconciliation")
        self.logger.info("=" * 80)

        self.load_multipliers(xtp_file=xtp_multipliers_file, jpm_file=jpm_multipliers_file)

        xtp_raw = self.load_position_file(xtp_file, 'XTP')
        jpm_raw = self.load_position_file(jpm_file, 'JPM')

        xtp_df = self._prepare_data(xtp_raw, 'XTP')
        jpm_df = self._prepare_data(jpm_raw, 'JPM')

        xtp_complete = xtp_df[xtp_df['_is_complete']].copy()
        xtp_incomplete = xtp_df[~xtp_df['_is_complete']].copy()
        jpm_complete = jpm_df[jpm_df['_is_complete']].copy()
        jpm_incomplete = jpm_df[~jpm_df['_is_complete']].copy()

        xtp_agg = self.aggregate_positions(xtp_complete, 'XTP') if len(xtp_complete) > 0 else pd.DataFrame()
        jpm_agg = self.aggregate_positions(jpm_complete, 'JPM') if len(jpm_complete) > 0 else pd.DataFrame()

        # Use multi-column merge keys (avoids string-concatenation collision risk)
        merge_keys = ['Account', 'Product_Code', 'Strike_Price_Adjusted', 'Settlement_Price_Adjusted']

        if len(xtp_agg) > 0 and len(jpm_agg) > 0:
            merged = pd.merge(
                xtp_agg, jpm_agg,
                on=merge_keys,
                how='outer',
                suffixes=('_XTP', '_JPM'),
            )
        elif len(xtp_agg) > 0:
            merged = xtp_agg.rename(columns={
                'Quantity': 'Quantity_XTP',
                'Strike_Price': 'Strike_Price_XTP',
                'Settlement_Price': 'Settlement_Price_XTP',
                'Strike_Multiplier': 'Strike_Multiplier_XTP',
                'Settle_Multiplier': 'Settle_Multiplier_XTP',
            })
            merged['Quantity_JPM'] = np.nan
        elif len(jpm_agg) > 0:
            merged = jpm_agg.rename(columns={
                'Quantity': 'Quantity_JPM',
                'Strike_Price': 'Strike_Price_JPM',
                'Settlement_Price': 'Settlement_Price_JPM',
                'Strike_Multiplier': 'Strike_Multiplier_JPM',
                'Settle_Multiplier': 'Settle_Multiplier_JPM',
            })
            merged['Quantity_XTP'] = np.nan
        else:
            merged = pd.DataFrame()

        if len(merged) > 0:
            # Track presence before fillna so 0-quantity positions are not misclassified
            merged['_in_xtp'] = merged['Quantity_XTP'].notna()
            merged['_in_jpm'] = merged['Quantity_JPM'].notna()
            merged['Quantity_XTP'] = merged['Quantity_XTP'].fillna(0)
            merged['Quantity_JPM'] = merged['Quantity_JPM'].fillna(0)
            merged['Quantity_Difference'] = merged['Quantity_XTP'] - merged['Quantity_JPM']

            def _get_status(row):
                if not row['_in_jpm']:
                    return 'XTP Only'
                if not row['_in_xtp']:
                    return 'JPM Only'
                diff = abs(row['Quantity_Difference'])
                if diff <= self.qty_tolerance:
                    return 'Matched'
                if self.qty_percent_tolerance > 0:
                    max_qty = max(abs(row['Quantity_XTP']), abs(row['Quantity_JPM']))
                    if max_qty > 0 and (diff / max_qty) <= self.qty_percent_tolerance:
                        return 'Matched'
                return 'Qty Mismatch'

            merged['Status'] = merged.apply(_get_status, axis=1)
            merged = merged.drop(columns=['_in_xtp', '_in_jpm'])

        results = {
            'merged': merged,
            'xtp_detail': xtp_complete,
            'jpm_detail': jpm_complete,
            'xtp_incomplete': xtp_incomplete,
            'jpm_incomplete': jpm_incomplete,
            'xtp_agg': xtp_agg,
            'jpm_agg': jpm_agg,
            'multipliers': self.multipliers,
            'products_with_default_multipliers': self.products_with_default_multipliers,
        }

        self.logger.info("Reconciliation completed")
        return results

    def generate_summary_stats(self, results: Dict) -> pd.DataFrame:
        """Generate summary statistics."""
        merged = results.get('merged', pd.DataFrame())
        xtp_incomplete_count = len(results.get('xtp_incomplete', pd.DataFrame()))
        jpm_incomplete_count = len(results.get('jpm_incomplete', pd.DataFrame()))

        def count_status(status):
            if len(merged) == 0 or 'Status' not in merged.columns:
                return 0
            return len(merged[merged['Status'] == status])

        matched_count = count_status('Matched')
        total = len(merged)

        stats = {
            'Metric': [
                'Total XTP Records (Complete)',
                'Total JPM Records (Complete)',
                'XTP Records with Missing Data',
                'JPM Records with Missing Data',
                'Total Unique Groups',
                'Matched Groups',
                'Quantity Mismatch Groups',
                'XTP Only Groups',
                'JPM Only Groups',
                'Overall Match Percentage',
                'Products Using Default Multipliers',
            ],
            'Value': [
                len(results.get('xtp_detail', pd.DataFrame())),
                len(results.get('jpm_detail', pd.DataFrame())),
                xtp_incomplete_count,
                jpm_incomplete_count,
                total,
                matched_count,
                count_status('Qty Mismatch'),
                count_status('XTP Only'),
                count_status('JPM Only'),
                f"{matched_count / total * 100:.2f}%" if total > 0 else "0%",
                len(results.get('products_with_default_multipliers', set())),
            ],
        }
        return pd.DataFrame(stats)

    def export_results(self, results: Dict, output_file: str,
                       summary_stats: Optional[pd.DataFrame] = None):
        """
        Export reconciliation results to Excel.

        Args:
            results: Reconciliation results dictionary
            output_file: Path to output Excel file
            summary_stats: Pre-computed summary stats (computed here if not provided)
        """
        self.logger.info(f"Exporting results to: {output_file}")

        if summary_stats is None:
            summary_stats = self.generate_summary_stats(results)

        merged = results.get('merged', pd.DataFrame())

        # --- Matched tab ---
        _MATCHED_COLS = [
            'Account', 'Product Code',
            'Strike Price (Original)', 'Strike Price (Adjusted)',
            'Settlement Price (Original)', 'Settlement Price (Adjusted)',
            'XTP Quantity', 'JPM Quantity', 'Quantity Difference', 'Match Status',
        ]
        matched_display = pd.DataFrame(columns=_MATCHED_COLS)
        if len(merged) > 0 and 'Status' in merged.columns:
            matched = merged[merged['Status'] == 'Matched'].copy()
            if len(matched) > 0:
                strike_orig = 'Strike_Price_XTP' if 'Strike_Price_XTP' in matched.columns else 'Strike_Price'
                settle_orig = 'Settlement_Price_XTP' if 'Settlement_Price_XTP' in matched.columns else 'Settlement_Price'
                matched_display = matched[[
                    'Account', 'Product_Code',
                    strike_orig, 'Strike_Price_Adjusted',
                    settle_orig, 'Settlement_Price_Adjusted',
                    'Quantity_XTP', 'Quantity_JPM', 'Quantity_Difference', 'Status',
                ]].copy()
                matched_display.columns = _MATCHED_COLS

        # --- Unmatched tab ---
        _UNMATCHED_COLS = [
            'Account', 'Product Code',
            'Strike Price (Adjusted)', 'Settlement Price (Adjusted)',
            'XTP Quantity', 'JPM Quantity', 'Difference', 'Status',
        ]
        unmatched = pd.DataFrame()          # always defined before use
        unmatched_display = pd.DataFrame(columns=_UNMATCHED_COLS)
        if len(merged) > 0 and 'Status' in merged.columns:
            unmatched = merged[merged['Status'] != 'Matched'].copy()
            if len(unmatched) > 0:
                unmatched['Abs_Difference'] = unmatched['Quantity_Difference'].abs()
                unmatched = unmatched.sort_values(
                    by=['Status', 'Abs_Difference'], ascending=[True, False]
                )
                unmatched_display = unmatched[[
                    'Account', 'Product_Code',
                    'Strike_Price_Adjusted', 'Settlement_Price_Adjusted',
                    'Quantity_XTP', 'Quantity_JPM', 'Quantity_Difference', 'Status',
                ]].copy()
                unmatched_display.columns = _UNMATCHED_COLS
                # Replace 0 with blank for readability (only safe after we've confirmed
                # zeros came from fillna, not from genuine zero-quantity positions)
                unmatched_display['XTP Quantity'] = unmatched_display['XTP Quantity'].replace(0, '')
                unmatched_display['JPM Quantity'] = unmatched_display['JPM Quantity'].replace(0, '')

        # --- Detail breakdown tab (merge-based; no list-in-cell) ---
        group_keys = ['Account', 'Product_Code', 'Strike_Price_Adjusted', 'Settlement_Price_Adjusted']
        detail_df = pd.DataFrame()
        if len(unmatched) > 0:
            unmatched_keys = unmatched[group_keys].drop_duplicates()
            detail_parts = []
            for src, src_detail in [
                ('XTP', results.get('xtp_detail', pd.DataFrame())),
                ('JPM', results.get('jpm_detail', pd.DataFrame())),
            ]:
                if len(src_detail) > 0 and all(k in src_detail.columns for k in group_keys):
                    matched_detail = src_detail.merge(unmatched_keys, on=group_keys, how='inner')
                    if len(matched_detail) > 0:
                        keep = [c for c in group_keys + ['Strike_Price', 'Settlement_Price', 'Quantity']
                                if c in matched_detail.columns]
                        part = matched_detail[keep].copy()
                        part['Source'] = src
                        detail_parts.append(part)
            if detail_parts:
                detail_df = pd.concat(detail_parts, ignore_index=True)

        # --- Incomplete records tab ---
        incomplete_parts = []
        for src, key in [('XTP', 'xtp_incomplete'), ('JPM', 'jpm_incomplete')]:
            inc = results.get(key, pd.DataFrame())
            if len(inc) > 0:
                disp_cols = ['Account', 'Product_Code', 'Strike_Price', 'Settlement_Price', 'Quantity']
                avail = [c for c in disp_cols if c in inc.columns]
                part = inc[avail].copy()
                part['Source'] = src
                incomplete_parts.append(part)

        if incomplete_parts:
            incomplete_df = pd.concat(incomplete_parts, ignore_index=True)
            incomplete_df.rename(columns={
                'Product_Code': 'Product Code',
                'Strike_Price': 'Strike Price',
                'Settlement_Price': 'Settlement Price',
            }, inplace=True)
        else:
            incomplete_df = pd.DataFrame(
                columns=['Account', 'Product Code', 'Strike Price', 'Settlement Price', 'Quantity', 'Source']
            )

        # --- Config audit tab ---
        mult = results.get('multipliers', {})
        all_products = sorted(set(
            list(mult.get('strike_xtp', {}).keys()) +
            list(mult.get('strike_jpm', {}).keys()) +
            list(mult.get('settle_xtp', {}).keys()) +
            list(mult.get('settle_jpm', {}).keys())
        ))
        config_used = pd.DataFrame({
            'Product_Code':    all_products,
            'Strike_Mult_XTP': [mult.get('strike_xtp', {}).get(p, 1.0) for p in all_products],
            'Strike_Mult_JPM': [mult.get('strike_jpm', {}).get(p, 1.0) for p in all_products],
            'Settle_Mult_XTP': [mult.get('settle_xtp', {}).get(p, 1.0) for p in all_products],
            'Settle_Mult_JPM': [mult.get('settle_jpm', {}).get(p, 1.0) for p in all_products],
        }) if all_products else pd.DataFrame(
            columns=['Product_Code', 'Strike_Mult_XTP', 'Strike_Mult_JPM',
                     'Settle_Mult_XTP', 'Settle_Mult_JPM']
        )

        # --- Write Excel ---
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            matched_display.to_excel(writer, sheet_name='Matched', index=False)
            unmatched_display.to_excel(writer, sheet_name='Unmatched_All', index=False)
            incomplete_df.to_excel(writer, sheet_name='Incomplete_Records', index=False)
            if len(detail_df) > 0:
                detail_df.to_excel(writer, sheet_name='Detail_Breakdown', index=False)
            summary_stats.to_excel(writer, sheet_name='Summary', index=False)
            config_used.to_excel(writer, sheet_name='Config_Used', index=False)

        # --- Apply Excel formatting ---
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font

            wb = load_workbook(output_file)

            if 'Unmatched_All' in wb.sheetnames:
                ws = wb['Unmatched_All']
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                # Locate columns by name, not hardcoded index
                col_index = {cell.value: i for i, cell in enumerate(ws[1])}
                status_idx   = col_index.get('Status')
                xtp_qty_idx  = col_index.get('XTP Quantity')
                jpm_qty_idx  = col_index.get('JPM Quantity')
                diff_idx     = col_index.get('Difference')

                if status_idx is not None:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        status = row[status_idx].value
                        if status == 'XTP Only' and xtp_qty_idx is not None:
                            row[xtp_qty_idx].fill = PatternFill(
                                start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        elif status == 'JPM Only' and jpm_qty_idx is not None:
                            row[jpm_qty_idx].fill = PatternFill(
                                start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        elif status == 'Qty Mismatch' and diff_idx is not None:
                            row[diff_idx].fill = PatternFill(
                                start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")

            wb.save(output_file)
        except Exception as e:
            self.logger.warning(f"Could not apply Excel formatting: {e}")

        self.logger.info(f"Results exported successfully to: {output_file}")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Position Reconciliation Tool: XTP vs JPM',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('--xtp-file', type=str, default='input/1.XTP.xlsx',
                        help='Path to XTP position file (default: input/1.XTP.xlsx)')
    parser.add_argument('--jpm-file', type=str, default='input/2.JPM.xlsx',
                        help='Path to JPM position file (default: input/2.JPM.xlsx)')
    parser.add_argument('--xtp-multipliers-file', type=str, default=None,
                        help='Path to XTP multipliers CSV (default: config/xtp_multipliers.csv). '
                             'Format: Product_Code, Strike_Multiplier, Settle_Multiplier')
    parser.add_argument('--jpm-multipliers-file', type=str, default=None,
                        help='Path to JPM multipliers CSV (default: config/jpm_multipliers.csv). '
                             'Format: Product_Code, Strike_Multiplier, Settle_Multiplier')
    parser.add_argument('--output-dir', type=str, default='output',
                        help='Output directory for results (default: output)')
    parser.add_argument('--config-dir', type=str, default='config',
                        help='Configuration directory (default: config)')
    parser.add_argument('--log-dir', type=str, default='logs',
                        help='Log directory (default: logs)')
    parser.add_argument('--strike-tolerance', type=float, default=0.01,
                        help='Tolerance for strike price comparisons (default: 0.01)')
    parser.add_argument('--qty-tolerance', type=float, default=0.0,
                        help='Absolute tolerance for quantity differences (default: 0.0)')
    parser.add_argument('--qty-percent-tolerance', type=float, default=0.0,
                        help='Fractional tolerance for quantity differences (default: 0.0)')

    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    tool = ReconciliationTool(
        config_dir=args.config_dir,
        log_dir=args.log_dir,
        strike_tolerance=args.strike_tolerance,
        qty_tolerance=args.qty_tolerance,
        qty_percent_tolerance=args.qty_percent_tolerance,
    )

    try:
        results = tool.reconcile(
            xtp_file=args.xtp_file,
            jpm_file=args.jpm_file,
            xtp_multipliers_file=args.xtp_multipliers_file,
            jpm_multipliers_file=args.jpm_multipliers_file,
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"reconciliation_{timestamp}.xlsx"

        # Compute once; reuse for both export and stdout print
        summary_stats = tool.generate_summary_stats(results)
        tool.export_results(results, str(output_file), summary_stats=summary_stats)

        print("\n" + "=" * 80)
        print("RECONCILIATION SUMMARY")
        print("=" * 80)
        print(summary_stats.to_string(index=False))
        print("\n" + "=" * 80)

        if results['products_with_default_multipliers']:
            print(f"\nWARNING: Products using default multipliers: "
                  f"{results['products_with_default_multipliers']}")
            print("Please update multiplier CSV files to include these products.")

        print(f"\nResults exported to: {output_file}")
        print("=" * 80)

    except Exception as e:
        tool.logger.error(f"Reconciliation failed: {e}", exc_info=True)
        print(f"ERROR: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
